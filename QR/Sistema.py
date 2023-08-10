import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl

cap = cv2.VideoCapture(0)

#Variables:
mañana = []
tarde = []
noche = []

def infhora ():
    inf = datetime.now()
    fecha = inf.strftime('%Y:%M:%D')
    hora = inf.strftime('%H:%M:%S')

    return hora, fecha

while True:
    ret, frame = cap.read()
    cv2.putText(frame, 'Localice el QR', (160, 80),cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
    cv2.rectangle(frame,(170, 100), (470, 400), (0, 255, 0), 2)

    hora, fecha = infhora()
    diasem = datetime.today().weekday()

    print (diasem)

    #AÑO/MES/DÍA
    a, me, d = fecha [0:4], fecha [5:7], fecha[8:10]

    #HORA/MINUTO/SEGUNDO
    h, m, s = int(hora[0:2]), int(hora[3:5]), int(hora[6:8])

    #Creacion de archivo
    nomar = str(a) + '-' + str(me) + '-' + str(d)
    texth = str(h) + '-' + str(m) + '-' + str(s)
    print(texth)
    print (texth)
    wb = xl.Workbook()

    #Lectura codigos QR
    for codes in decode(frame):
        info = codes.data.decode('utf-8')

        tipo = info[0:2]
        tipo = int(tipo)
        letr = chr(tipo)

        num = info [2:]

        #extraccion de coordenadas:
        pts = np.array([codes.polygon], np.int32)
        xi, yi = codes.rect.left, codes.rect.top

        pts = pts.reshape((-1, 1, 2))

        #ID COMPLETO
        codigo = letr + num

        #SEMANA
        #MAÑANA
        if 4 >= diasem >= 0:

            if 10 >= h >= 7:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                #Guardar id:
                if codigo not in mañana:
                    #Agregar id:
                    pos = len(mañana)
                    mañana.append(codigo)

                    hojam = wb.create_sheet ("Mañana")
                    datos = hojam.append(mañana)
                    wb.save(nomar + '.xlsx')

                    cv2.putText(frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                elif codigo in mañana:
                    cv2.putText (frame, 'El ID ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)

            #TARDE   
            if 14 >= h >= 11:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in tarde:
                    tarde.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    
                    hojat = wb.create_sheet ("Tarde")
                    datos = hojat.append(tarde)
                    wb.save(nomar + '.xlsx')

                elif codigo in tarde:
                    cv2.putText (frame, 'El ID ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)
                
                print (tarde)
                    
            #NOCHE    
            if 20 >= h >= 11:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in noche:
                    noche.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    
                    hojan = wb.create_sheet ("Noche")
                    datos = hojan.append(tarde)
                    wb.save(nomar + '.xlsx')

                elif codigo in noche:
                    cv2.putText (frame, 'El ID ' + str(codigo), 
                                 (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                 (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)

                print(noche)

        #FIN DE SEMANA
        #MAÑANA
        if 6 >= diasem >= 5:

            if 8 >= h >=6:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in mañana:
                    mañana.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                    hojam = wb.create_sheet ("Mañana")
                    datos = hojam.append(mañana)
                    wb.save(nomar + '.xlsx')

                elif codigo in mañana:
                    cv2.putText (frame, 'El roster ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)
                    
                print (mañana)

            #TARDE
            if h == 11 and m >= 30:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in tarde:
                    tarde.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    
                    hojat = wb.create_sheet ("Tarde")
                    datos = hojat.append(tarde)
                    wb.save(nomar + '.xlsx')
                elif codigo in tarde:
                    cv2.putText (frame, 'El ID ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)
                
            elif h == 13 and m <= 30: 
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)

                if codigo not in tarde:
                    mañana.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                    #hojam = wb.create_sheet ("Mañana")
                    datos = hojat.append(tarde)
                    wb.save(nomar + '.xlsx')

                elif codigo in tarde:
                    cv2.putText (frame, 'El roster ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)
            
            #NOCHE
            if 19 >= h >=17:
                cv2.polylines(frame, [pts], True, (255, 255, 0), 5)
                if codigo not in noche:
                    noche.append(codigo)
                    cv2.putText (frame, letr + '0' + str(num), (xi - 15, yi -15), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)

                    hojan = wb.create_sheet ("Noche")
                    datos = hojan.append(noche)
                    wb.save(nomar + '.xlsx')

                elif codigo in noche:
                    cv2.putText (frame, 'El ID ' + str(codigo),
                                (xi - 65, yi - 45), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 55, 0), 2)
                    cv2.putText(frame, 'Fue registrado',
                                (xi - 65, yi - 15), cv2.FONT_HERSHEY_SIMPLEX, 1 (255, 0, 0), 2)
    
    cv2.imshow(" QR READER ", frame)

    t=cv2.waitKey(5)
    if t == 27:
        break

cv2.destroyAllWindows()
cap.release()